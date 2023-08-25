# import general known functions
# import specific functions just used in this file
import csv
import logging
import sys
from numbers import Number
from pathlib import Path

import openpyxl

from base_python.source.helper._FunctionDef import *


def ReadTable_sheet(filename):
    """Read some table from *.csv or *.xlsx file"""

    myrows = []

    # check if it is a csv and open as textfile
    if filename.find('.csv') > 0:
        with open(filename, newline='') as csvfile:
            lines = csv.reader(csvfile, delimiter=';', quotechar='"')
            line_index = 0
            row_max = -1
            for line in lines:
                row_index = 0
                for value in line:  # convert text to numbers
                    if row_index > row_max:
                        row_max = row_index
                        myrows.append([])
                    try:
                        locale.atof(value)
                    except Exception:
                        if len(value) > 0:
                            myrows[row_index].append(value)
                        pass
                    else:
                        myrows[row_index].append(locale.atof(value))
                    row_index += 1

                line_index += 1

        return myrows

    # check if it is a xlsx and open with openpyxl
    elif filename.find('.xlsx') > 0:
        xlsx_file = Path(filename)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        wb_obj.active = 0  # set the first sheet as active
        sheet = wb_obj.active  # Read the active sheet

        for i in range(sheet.max_column):
            line = []
            for row in sheet.iter_rows(1, sheet.max_row):
                if row[i].stream is not None:
                    line.append(row[i].stream)
            myrows.append(line)

        return myrows

    else:
        logging.critical('ReadTable: the specified file does not match the supported file formats: *.csv, *.xlsx')
        sys.exit()


def ReadTable_dataset(filename, rowindex):
    """Read some dataset as vector from *.csv or *.xlsx file; specify the desired row"""

    # import the file using the above function
    data = ReadTable_sheet(filename)
    myrow = []
    row_header = ''

    # check whether the line is a number, if not ignore and through a message
    line_index = -1
    for line in data[rowindex]:
        line_index += 1
        if isinstance(line, Number):
            myrow.append(line)
        elif line is not None:
            logging.debug(
                'in file "' + filename + '" at line {}: "{}" is not an number and will be ignored'.format(line_index,
                                                                                                          line))
            row_header = line

    row_header = row_header.split()[0]

    return [myrow, row_header]


def ReadTable_by_columns(filename):
    """Read some dataset as vector from *.csv or *.xlsx file; return al columns as dictionary of the first row"""

    # import the file using the above function
    data = ReadTable_sheet(filename)
    data_dict = {}

    for column in data:
        if len(column) > 0:
            data_dict[str(column[0])] = column[1:]

    return data_dict
